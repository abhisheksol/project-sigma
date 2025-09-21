# utils/export_field_mapping.py
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Callable
from store.configurations.loan_config.template_config.models import (
    ProcessTemplateFieldMappingModel,
    ProcessTemplateMultiReferenceFieldModel,
)
from django.db.models.query import QuerySet
from django.core.management.base import BaseCommand


def static_export_field_mapping_headers(file_path: str):
    """
    Export only the column headers (labels) from ProcessTemplateFieldMappingModel into Excel,
    sorted by ordering.
    """
    queryset: QuerySet[ProcessTemplateFieldMappingModel] = (
        ProcessTemplateFieldMappingModel.objects.all()
        .order_by("ordering")  # ✅ sort by ordering
        .values("label", "title", "ordering")
    )

    if not queryset.exists():
        raise ValueError("No data found in ProcessTemplateFieldMappingModel")

    # Extract headers sorted by ordering (fallback to title if label missing)
    headers: List[str] = [item["label"] or item["title"] for item in queryset]

    wb: openpyxl.Workbook = openpyxl.Workbook()
    ws: Callable = wb.active
    ws.title = "Headers"

    # Write headers in first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter: str = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(15, len(str(header)) + 2)

    wb.save(file_path)
    return file_path


def dynamic_export_field_mapping_headers(file_path: str):
    """
    Export headers including dynamic (multi-ref) fields into Excel.
    - Normal fields -> single header
    - Dynamic fields -> expand into multiple headers (child labels)
    """
    queryset: QuerySet[ProcessTemplateFieldMappingModel] = (
        ProcessTemplateFieldMappingModel.objects.all()
        .order_by("ordering")
        .values("id", "label", "title", "ordering", "is_multi_ref_field")
    )

    if not queryset.exists():
        raise ValueError("No data found in ProcessTemplateFieldMappingModel")

    headers: List = []
    for item in queryset:
        if item["is_multi_ref_field"]:
            # Expand into child labels
            children = ProcessTemplateMultiReferenceFieldModel.objects.filter(
                reference_field_id=item["id"]
            ).values("label", "title")

            if children.exists():
                headers.extend([c["label"] or c["title"] for c in children])
            else:
                # fallback to parent label if no children
                headers.append(item["label"] or item["title"])
        else:
            headers.append(item["label"] or item["title"])

    wb: openpyxl.Workbook = openpyxl.Workbook()
    ws: Callable = wb.active
    ws.title = "Headers"

    # Write headers in first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(15, len(str(header)) + 2)

    wb.save(file_path)
    return file_path


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="field_mappings.xlsx",
            help="Path to save the Excel file (default: field_mappings.xlsx)",
        )

    def handle(self, *args, **options):
        file_path: str = options["file"]
        try:
            path = static_export_field_mapping_headers(file_path)
            self.stdout.write(self.style.SUCCESS(f"Export successful → {path}"))
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Export failed: {str(e)}"))
